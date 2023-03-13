import requests
from bs4 import BeautifulSoup
from time import sleep
import re
import openpyxl
from PySimpleGUI import PySimpleGUI as sg

#  Definindo tema
sg.theme('Reddit')
layout = [
    [sg.Text('Pesquisa ')],
    [sg.Input()],
    [sg.Text('1 página = 10 pessoas')],
    [sg.Text('Quantidade:')],
    [sg.Input()],
    [sg.Button('Buscar')]
]

#Justando janela do PySimpleGUI
window = sg.Window('Mercado Livre', layout)


def scraping ():

    #compra = (valores[0])

    sleep(3)

    url_base = 'https://lista.mercadolivre.com.br/'
    url = (url_base+compra)

    #excel:
    wb = openpyxl.Workbook()
    ws = wb.active
    class Buscar():
        
        def get_text(valor_default=compra):
            
            #realizando o get da pagina
            resposta = requests.get(url)
        
            #ativando o BeautifulSoup, para "Arrumar" o codigo HTML
            site = BeautifulSoup(resposta.text, 'html.parser')
        
            #div com o produto
            produtos = site.findAll('div', attrs={'class': 'ui-search-result__wrapper shops__result-wrapper'})
            count = 0
            for produto in produtos:

                count+=1
                #titulo do produto, com o nome
                #erro ao imprimir
                titulo = produto.find('h2', attrs= {'class': 'ui-search-item__title'})
                #titulo2 = produto.find('div', class_ = re.compile('ui-search-item__title'))
                
                #selecionando apenas o valor em rais
                reais = produto.find('span', attrs={'class': 'price-tag-text-sr-only'})
                #centavos = produto.find('span', attrs={'class': ''})
            
                #selecionando o link em uma variavel(link)
                link = produto.find('a', attrs={'class':'ui-search-link'})

                print('Contador: ', count)
                print('Produto:', titulo.text)
                print('Valor do produto: R$',reais.text )
                print('Link da pagina:',link['href'])

                ws['A1'] = 'TITULO'
                ws['B1'] = 'VALOR EM REAL'

                for i in range(0,count):
                    i = count+1
                    ws[f'A{i}'] = titulo.text
                    ws[f'B{i}'] = reais.text

                    wb.save('data.xlsx')

                print('\n\n')
        
        get_text()
        

    Buscar()
    
#Ler os eventos do front end
while True:
    
    
    eventos, valores = window.read()
    #var do input do Layout
    compra = (valores[0])
    # input de quantas paginas
    #input_page = 1
    print(valores[0])
    input_page = (valores[1])  
    
    if eventos == sg.WINDOW_CLOSED:
        
        print('fechando...')
        break
        
    if eventos == 'Buscar':
        scraping()
        print('tudo correto.')
        break
